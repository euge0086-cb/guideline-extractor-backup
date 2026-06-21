"""
GUIDELINE REFERENCE PIPELINE
================================
Pipeline modular para extraer referencias de guías clínicas en PDF,
enriquecer con metadatos (PubMed + CrossRef) y exportar a Excel.

MÓDULOS:
  1. extract_references(pdf_path)  → lista de strings de referencias crudas
  2. enrich_reference(ref_text)    → dict con PMID, DOI, autores, año, etc.
  3. classify_reference(metadata)  → tipo: RCT_primario / RCT_secundario / meta-analisis / observacional / otro
  4. export_to_excel(records, out) → archivo .xlsx con hojas diferenciadas

USO:
  python guideline_pipeline.py <ruta_al_pdf> [output.xlsx]
"""

import re
import sys
import time
import json
import requests
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# MÓDULO 1: EXTRACCIÓN DE REFERENCIAS DEL PDF
# ─────────────────────────────────────────────

def extract_references_from_pdf(pdf_path: str, debug: bool = False) -> list[str]:
    """
    Extrae el bloque de referencias de un PDF de guía clínica.
    Soporta PDFs de una y dos columnas, mayúsculas/minúsculas,
    y formatos con numeración de página intercalada.

    Estrategia de 3 niveles:
      A. Detección por columnas con header de sección "References"
      B. Fallback: texto completo lineal con header de sección
      C. Fallback final: buscar el ÚLTIMO bloque de líneas numeradas
         consecutivas del documento (las guías casi siempre acaban así)
    """
    log = []

    def dbg(msg):
        log.append(msg)
        if debug:
            print(msg)

    # Patrones de header de sección "References" (tolerante a mayúsculas,
    # números de página pegados, dos puntos, etc.)
    HEADER_PATTERNS = [
        r'^references?\s*:?\s*\d*\s*$',
        r'^bibliograf[ií]a\s*:?\s*\d*\s*$',
        r'^referenci[ae]s\s*:?\s*\d*\s*$',
        r'^literature\s+cited\s*$',
        r'^works\s+cited\s*$',
    ]

    def is_header(line: str) -> bool:
        line_clean = line.strip()
        for pat in HEADER_PATTERNS:
            if re.match(pat, line_clean, re.IGNORECASE):
                return True
        return False

    def is_ref_start(line: str) -> bool:
        """Detecta inicio de referencia numerada: '1.', '12.', '1)', etc."""
        return bool(re.match(r'^\d{1,4}[\.\)]\s+\S', line.strip()))

    # ── NIVEL A: extracción por columnas con detección de header ──────────
    all_lines_by_page = []
    header_found_at = None

    with pdfplumber.open(pdf_path) as pdf:
        n_pages = len(pdf.pages)
        dbg(f"[INFO] PDF con {n_pages} páginas")

        for page_idx, page in enumerate(pdf.pages):
            page_width = page.width
            page_lines = []

            left_col = page.crop((0, 0, page_width / 2, page.height))
            right_col = page.crop((page_width / 2, 0, page_width, page.height))

            for col in [left_col, right_col]:
                try:
                    words = col.extract_words()
                except Exception:
                    words = []
                if not words:
                    continue
                line_map = {}
                for w in words:
                    y_key = round(w["top"])
                    line_map.setdefault(y_key, []).append(w["text"])
                for y_key in sorted(line_map.keys()):
                    line = " ".join(line_map[y_key]).strip()
                    if line:
                        page_lines.append(line)

            all_lines_by_page.append(page_lines)

            if header_found_at is None:
                for line in page_lines:
                    if is_header(line):
                        header_found_at = page_idx
                        dbg(f"[INFO] Header 'References' encontrado en página {page_idx + 1}: '{line}'")
                        break

    if header_found_at is not None:
        # Reunir todas las líneas desde el header en adelante
        all_lines = []
        collecting = False
        for page_idx, page_lines in enumerate(all_lines_by_page):
            for line in page_lines:
                if not collecting:
                    if is_header(line):
                        collecting = True
                    continue
                all_lines.append(line)
        dbg(f"[INFO] Nivel A (columnas): {len(all_lines)} líneas recogidas tras el header")
    else:
        all_lines = []
        dbg("[WARN] Nivel A falló: no se encontró header 'References' por columnas")

    # ── NIVEL B: fallback texto lineal simple (sin separar columnas) ──────
    if not all_lines:
        dbg("[INFO] Probando Nivel B: extracción de texto simple")
        full_text = ""
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                t = page.extract_text()
                if t:
                    full_text += t + "\n"

        header_pos = None
        for pat in HEADER_PATTERNS:
            m = re.search(pat, full_text, re.IGNORECASE | re.MULTILINE)
            if m:
                header_pos = m.end()
                dbg(f"[INFO] Nivel B: header encontrado con patrón '{pat}'")
                break

        if header_pos:
            all_lines = full_text[header_pos:].split("\n")
            dbg(f"[INFO] Nivel B: {len(all_lines)} líneas tras el header")

    # ── NIVEL C: último recurso — buscar el último bloque numerado denso ──
    if not all_lines:
        dbg("[INFO] Probando Nivel C: detección del último bloque numerado")
        # Concatenar todas las líneas de todas las páginas en orden
        flat_lines = []
        for page_lines in all_lines_by_page:
            flat_lines.extend(page_lines)

        # Buscar posiciones de líneas que parecen inicio de referencia
        ref_start_indices = [i for i, l in enumerate(flat_lines) if is_ref_start(l)]

        if len(ref_start_indices) >= 5:
            # Verificar que las referencias siguen una numeración creciente
            # consecutiva (tolerando algún salto pequeño por ruido de OCR)
            numbers = []
            for i in ref_start_indices:
                m = re.match(r'^(\d{1,4})', flat_lines[i].strip())
                if m:
                    numbers.append((i, int(m.group(1))))

            # Encontrar el tramo más largo de numeración creciente
            best_start = None
            best_len = 0
            run_start = 0
            for k in range(1, len(numbers)):
                if numbers[k][1] <= numbers[k-1][1] or numbers[k][1] - numbers[k-1][1] > 3:
                    run_len = k - run_start
                    if run_len > best_len:
                        best_len = run_len
                        best_start = run_start
                    run_start = k
            run_len = len(numbers) - run_start
            if run_len > best_len:
                best_len = run_len
                best_start = run_start

            if best_start is not None and best_len >= 5:
                start_line_idx = numbers[best_start][0]
                all_lines = flat_lines[start_line_idx:]
                dbg(f"[INFO] Nivel C: bloque numerado encontrado desde línea {start_line_idx}, "
                    f"{best_len} referencias detectadas en la secuencia")

    if not all_lines:
        dbg("[ERROR] Los 3 niveles de extracción fallaron. "
            "El PDF puede tener un formato no soportado (escaneado, protegido, o sin referencias numeradas).")
        if not debug:
            for l in log:
                print(l)
        return []

    # ── Unir líneas en referencias completas ───────────────────────────────
    full_refs = []
    current = ""
    page_footer_pattern = re.compile(
        r'^(Representativeness|European|Eur\s+Heart|Circulation|JAMA|J\s+Am\s+Coll|'
        r'Downloaded\s+from|page\s+\d+|^\d{1,4}$)', re.IGNORECASE
    )

    for line in all_lines:
        line = line.strip()
        if not line:
            continue
        if is_ref_start(line):
            if current:
                full_refs.append(re.sub(r'\s+', ' ', current).strip())
            current = line
        elif current:
            if page_footer_pattern.match(line):
                continue
            current += " " + line

    if current:
        full_refs.append(re.sub(r'\s+', ' ', current).strip())

    full_refs = [r for r in full_refs if len(r) > 25]

    dbg(f"[INFO] Referencias extraídas (total final): {len(full_refs)}")
    if not debug:
        for l in log:
            print(l)

    return full_refs


# ─────────────────────────────────────────────
# MÓDULO 2: ENRIQUECIMIENTO VÍA APIs
# ─────────────────────────────────────────────

def search_pubmed(query: str, api_key: str = "") -> dict:
    """Busca en PubMed: PMID + metadatos + Publication Types oficiales."""
    base = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/"
    params = {"db": "pubmed", "term": query, "retmax": 1, "retmode": "json"}
    if api_key: params["api_key"] = api_key
    try:
        r = requests.get(f"{base}esearch.fcgi", params=params, timeout=10)
        ids = r.json().get("esearchresult", {}).get("idlist", [])
        if not ids: return {}
        pmid = ids[0]
        fetch_params = {"db": "pubmed", "id": pmid, "retmode": "xml", "rettype": "abstract"}
        if api_key: fetch_params["api_key"] = api_key
        rf = requests.get(f"{base}efetch.fcgi", params=fetch_params, timeout=10)
        xml = rf.text
        def extract_xml(tag, text):
            m = re.search(rf'<{tag}[^>]*>(.*?)</{tag}>', text, re.DOTALL)
            return m.group(1).strip() if m else ""
        title   = extract_xml("ArticleTitle", xml)
        year    = extract_xml("Year", xml) or extract_xml("MedlineDate", xml)[:4]
        journal = extract_xml("Title", xml)
        authors_raw = re.findall(r'<LastName>(.*?)</LastName>.*?<ForeName>(.*?)</ForeName>', xml, re.DOTALL)
        authors = ", ".join([f"{ln} {fn[0]}." for ln, fn in authors_raw[:3]])
        if len(authors_raw) > 3: authors += " et al."
        doi_m = re.search(r'<ArticleId IdType="doi">(.*?)</ArticleId>', xml)
        doi = doi_m.group(1).strip() if doi_m else ""
        pub_types  = re.findall(r'<PublicationType[^>]*>(.*?)</PublicationType>', xml)
        mesh_terms = re.findall(r'<DescriptorName[^>]*>(.*?)</DescriptorName>', xml)
        return {"pmid": pmid, "doi": doi, "title": title,
                "year": year[:4] if year else "", "journal": journal,
                "authors": authors, "pub_types": pub_types,
                "mesh_terms": mesh_terms, "source": "PubMed"}
    except:
        return {}


def fetch_pubtypes_by_pmid(pmid: str, api_key: str = "") -> tuple:
    """Obtiene Publication Types y MeSH de PubMed dado un PMID ya conocido."""
    base = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi"
    params = {"db": "pubmed", "id": pmid, "retmode": "xml", "rettype": "abstract"}
    if api_key: params["api_key"] = api_key
    try:
        r = requests.get(base, params=params, timeout=10)
        pub_types  = re.findall(r'<PublicationType[^>]*>(.*?)</PublicationType>', r.text)
        mesh_terms = re.findall(r'<DescriptorName[^>]*>(.*?)</DescriptorName>', r.text)
        return pub_types, mesh_terms
    except:
        return [], []


def search_crossref(ref_text: str) -> dict:
    """Busca en CrossRef por texto de referencia libre."""
    url = "https://api.crossref.org/works"
    # Limpiar número de referencia
    clean = re.sub(r'^\d+[\.\s]+', '', ref_text).strip()[:200]
    params = {
        "query.bibliographic": clean,
        "rows": 1,
        "select": "DOI,title,author,published,container-title,type"
    }
    headers = {"User-Agent": "Guideline-Pipeline/1.0 (research tool)"}
    try:
        r = requests.get(url, params=params, headers=headers, timeout=10)
        items = r.json().get("message", {}).get("items", [])
        if not items:
            return {}
        item = items[0]
        
        doi = item.get("DOI", "")
        title = item.get("title", [""])[0] if item.get("title") else ""
        pub_type = item.get("type", "")
        
        authors_list = item.get("author", [])
        authors = ", ".join([
            f"{a.get('family', '')} {a.get('given', [''])[0]}." 
            if a.get('given') else a.get('family', '')
            for a in authors_list[:3]
        ])
        if len(authors_list) > 3:
            authors += " et al."
        
        year = ""
        pub = item.get("published", {}).get("date-parts", [[""]])
        if pub and pub[0]:
            year = str(pub[0][0])
        
        journal = ""
        ct = item.get("container-title", [])
        if ct:
            journal = ct[0]
        
        return {
            "doi": doi,
            "title": title,
            "year": year,
            "journal": journal,
            "authors": authors,
            "pub_type_raw": pub_type,
            "source": "CrossRef"
        }
    except Exception as e:
        return {}


def build_pubmed_query(ref_text: str) -> str:
    """Construye query PubMed desde texto de referencia."""
    # Extraer título tentativo: texto entre el primer punto y la revista
    clean = re.sub(r'^\d+[\.\s]+', '', ref_text).strip()
    # Intentar extraer primeras palabras significativas del título
    words = re.findall(r'\b[A-Za-z]{4,}\b', clean)
    query = " ".join(words[:8])
    # Añadir año si está presente
    year_m = re.search(r'\b(19|20)\d{2}\b', clean)
    if year_m:
        query += f"[Title/Abstract] AND {year_m.group()}[PDAT]"
    return query


def enrich_reference(ref_text: str, idx: int) -> dict:
    """
    Enriquece una referencia con metadatos desde PubMed y CrossRef.
    Combina los resultados priorizando PubMed para PMID y CrossRef para DOI.
    """
    record = {
        "ref_number": idx,
        "ref_raw": ref_text,
        "pmid": "",
        "doi": "",
        "title": "",
        "authors": "",
        "year": "",
        "journal": "",
        "study_type": "",
        "study_type_auto": "",
        "pubmed_url": "",
        "doi_url": "",
        "source_api": "",
        "pub_types": [],    # Publication Types oficiales de PubMed
        "mesh_terms": [],   # MeSH terms de PubMed
        "pub_type_raw": "", # Tipo CrossRef
        "notes": ""
    }

    time.sleep(0.35)  # Respetar rate limit NCBI (3 req/s sin API key)

    # 1. CrossRef primero (más tolerante a texto libre)
    cr = search_crossref(ref_text)
    if cr:
        for k, v in cr.items():
            if v and k in record:
                record[k] = v
        record["source_api"] = "CrossRef"

    # 2. PubMed para PMID + Publication Types (clave para clasificación)
    query = build_pubmed_query(ref_text)
    if query:
        time.sleep(0.35)
        pm = search_pubmed(query)
        if pm:
            if pm.get("pmid"):
                record["pmid"] = pm["pmid"]
                record["pubmed_url"] = f"https://pubmed.ncbi.nlm.nih.gov/{pm['pmid']}/"
            for k in ["doi", "title", "authors", "year", "journal"]:
                if not record[k] and pm.get(k):
                    record[k] = pm[k]
            # Guardar Publication Types y MeSH para el clasificador
            record["pub_types"]  = pm.get("pub_types", [])
            record["mesh_terms"] = pm.get("mesh_terms", [])
            record["source_api"] = "PubMed+CrossRef" if cr else "PubMed"

    # 3. Si ya tenemos PMID pero no pub_types (vino solo de CrossRef), buscarlos
    if record["pmid"] and not record["pub_types"]:
        time.sleep(0.35)
        pt, mt = fetch_pubtypes_by_pmid(record["pmid"])
        record["pub_types"]  = pt
        record["mesh_terms"] = mt

    if record["doi"]:
        record["doi_url"] = f"https://doi.org/{record['doi']}"

    return record


# ─────────────────────────────────────────────
# MÓDULO 3: CLASIFICACIÓN AUTOMÁTICA
# ─────────────────────────────────────────────

# Palabras clave para clasificación por tipo de estudio
CLASSIFICATION_RULES = {
    "RCT_primario": [
        r'\brandomis[ei]d\b', r'\bplacebo.controlled\b',
        r'\bdouble.blind\b', r'\bsingle.blind\b',
        r'\brandom(ized|ised)\s+(clinical|controlled)\s+trial\b',
        r'\bRCT\b', r'\bensayo\s+cl[ií]nico\b', r'\brandomizado\b',
        r'\bprimary\s+(result|endpoint|outcome)\b',
        # Nombres de trials clásicos en texto crudo
        r'\bPURSUIT\b', r'\bPRISM\b', r'\bTIMI\s+III\b', r'\bGUSTO\b',
        r'\bCURE\b', r'\bACUITY\b', r'\bTRILOGY\b', r'\bPLATO\b',
        r'\bBRILLIANT\b', r'\bTARGET\b', r'\bSYNERGY\b',
        r'\btrial\b.*\bplacebo\b', r'\bversus\b.*\bplacebo\b',
    ],
    "RCT_secundario": [
        r'\bsubgroup\s+anal', r'\bpost.hoc\b', r'\bsecondary\s+anal',
        r'\bsub-?study\b', r'\bpre.specified\b', r'\bpost\s+hoc\b',
        r'\bsubstudy\b', r'\bsub\s+analysis\b',
    ],
    "meta-analisis": [
        r'\bmeta.anal', r'\bsystematic\s+review\b', r'\bpooled\s+anal',
        r'\bsystematic\b.*\breview\b', r'\bmetaan[aá]lisis\b',
        r'\bindividual\s+patient\s+data\b', r'\bnetwork\s+meta\b',
    ],
    "registro_observacional": [
        r'\bregist(ry|er|ro)\b', r'\bcohort\b', r'\bobservational\b',
        r'\bretrospective\b', r'\bprospective\s+(cohort|observational)\b',
        r'\bepidemiolog\b', r'\bsurvey\b', r'\bpopulation.based\b',
        r'\bdatabase\b', r'\bcross.sectional\b',
        # Registros conocidos en texto crudo
        r'\bGRACE\b', r'\bSWEDEHEART\b', r'\bNRMI\b', r'\bCRUSADE\b',
        r'\bACTION\b.*\bregist', r'\bEHS\b', r'\bEURO\s*HEART\b',
        r'\bNHANES\b', r'\bFRAMINGHAM\b',
    ],
    "guia_clinica": [
        r'\bguidelines?\b', r'\brecommendations?\b',
        r'\bconsensus\s+(statement|document|report)\b',
        r'\bgu[ií]a\s+(cl[ií]nica|de\s+pr[aá]ctica)\b',
        r'\btask\s+force\b', r'\bwriting\s+(committee|group)\b',
        r'\bpractice\s+guideline\b', r'\bexpert\s+consensus\b',
        r'\bposition\s+(statement|paper)\b',
        r'\bESC\s+guideline', r'\bACC.AHA\s+guideline',
        r'\bACCF.AHA\s+guideline', r'\bAHA.ACC\s+guideline',
        r'\bfocused\s+update\b', r'\bpolicy\s+statement\b',
        r'\bscientific\s+statement\b',
    ],
}

# Mapeo de Publication Types de PubMed → categorías de la app
PUBMED_TYPE_MAP = {
    # ECA primario
    "Randomized Controlled Trial":              "RCT_primario",
    "Controlled Clinical Trial":                "RCT_primario",
    "Clinical Trial, Phase III":                "RCT_primario",
    "Clinical Trial, Phase IV":                 "RCT_primario",
    "Multicenter Study":                        None,  # complementario, no definitivo
    # ECA secundario
    "Clinical Trial":                           None,  # demasiado genérico solo
    # Meta-análisis
    "Meta-Analysis":                            "meta-analisis",
    "Systematic Review":                        "meta-analisis",
    # Observacional / registro
    "Observational Study":                      "registro_observacional",
    "Multicenter Study":                        None,
    # Guía clínica
    "Practice Guideline":                       "guia_clinica",
    "Guideline":                                "guia_clinica",
    "Consensus Development Conference":         "guia_clinica",
    "Consensus Development Conference, NIH":    "guia_clinica",
    "Government Publications":                  None,
}

def classify_reference(record: dict) -> str:
    """
    Clasificación en 3 capas de precisión decreciente:
    1. Publication Types oficiales de PubMed (más fiable)
    2. Palabras clave en título + texto crudo + journal
    3. Fallback por tipo CrossRef + vocabulario de intervención
    """

    # ── CAPA 1: Publication Types de PubMed ─────────────────────────────
    pub_types = record.get("pub_types", [])  # lista de strings de PubMed
    if pub_types:
        # Prioridad: RCT_secundario se detecta por combinación de tipos
        type_str = " | ".join(pub_types)

        # Subanálisis: Clinical Trial + sin "Randomized" = probable secundario
        if any("Randomized" in t for t in pub_types):
            if any(kw in type_str for kw in ["Subgroup", "Secondary", "Post-Hoc"]):
                return "RCT_secundario"

        # Meta-análisis y revisiones sistemáticas
        if any(t in ("Meta-Analysis", "Systematic Review") for t in pub_types):
            return "meta-analisis"

        # Guías clínicas
        if any(t in ("Practice Guideline", "Guideline",
                     "Consensus Development Conference",
                     "Consensus Development Conference, NIH") for t in pub_types):
            return "guia_clinica"

        # ECA primario
        if any(t == "Randomized Controlled Trial" for t in pub_types):
            return "RCT_primario"

        # Observacional
        if any(t in ("Observational Study",) for t in pub_types):
            return "registro_observacional"

    # ── CAPA 2: Palabras clave en todos los campos de texto ──────────────
    text_to_search = " ".join(filter(None, [
        record.get("title", ""),
        record.get("ref_raw", ""),
        record.get("pub_type_raw", ""),
        record.get("journal", ""),
        " ".join(record.get("mesh_terms", [])),
    ]))

    for study_type in ["RCT_secundario", "meta-analisis", "guia_clinica",
                        "registro_observacional", "RCT_primario"]:
        for pat in CLASSIFICATION_RULES[study_type]:
            if re.search(pat, text_to_search, re.IGNORECASE):
                return study_type

    # ── CAPA 3: Fallback por vocabulario de intervención ─────────────────
    if record.get("pub_type_raw") == "journal-article":
        title = record.get("title", "")
        for pat in [r'\beffect\s+of\b', r'\befficacy\b',
                    r'\bsafety\s+and\s+efficacy\b', r'\bversus\b',
                    r'\bcompar(ing|ison)\b', r'\bbenefit\s+of\b']:
            if re.search(pat, title, re.IGNORECASE):
                return "RCT_primario"

    return "otro/no_clasificado"


# ─────────────────────────────────────────────
# MÓDULO 4: EXPORTACIÓN A EXCEL
# ─────────────────────────────────────────────

COLORS = {
    "header_bg": "1F4E79",
    "header_fg": "FFFFFF",
    "rct_primary": "E2EFDA",    # verde claro
    "rct_secondary": "FFF2CC",  # amarillo claro
    "meta": "DAE8FC",           # azul claro
    "registry": "F8CECC",       # rosa claro
    "guideline": "E1D5E7",      # lila claro
    "other": "F5F5F5",          # gris claro
    "subheader": "BDD7EE",
}

STUDY_TYPE_COLORS = {
    "RCT_primario": COLORS["rct_primary"],
    "RCT_secundario": COLORS["rct_secondary"],
    "meta-analisis": COLORS["meta"],
    "registro_observacional": COLORS["registry"],
    "guia_clinica": COLORS["guideline"],
    "otro/no_clasificado": COLORS["other"],
}

def style_header(cell, bg_color=None, fg_color="FFFFFF", bold=True):
    bg = bg_color or COLORS["header_bg"]
    cell.font = Font(bold=bold, color=fg_color, name="Arial", size=10)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def style_cell(cell, row_color=None, wrap=False):
    cell.font = Font(name="Arial", size=9)
    if row_color:
        cell.fill = PatternFill("solid", start_color=row_color)
    cell.alignment = Alignment(vertical="top", wrap_text=wrap)

def add_thin_border(ws, row, col_start, col_end):
    thin = Side(style="thin", color="CCCCCC")
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = Border(bottom=thin)


def export_to_excel(records: list[dict], output_path: str):
    wb = Workbook()

    # ── Hoja 1: Base de datos completa ──────────────────────────────
    ws_all = wb.active
    ws_all.title = "Todas las referencias"

    columns = [
        ("N°", 5), ("Autores", 30), ("Año", 6), ("Título", 50),
        ("Revista", 25), ("PMID", 12), ("DOI", 30),
        ("URL PubMed", 35), ("URL DOI", 35),
        ("Tipo (auto)", 20), ("Tipo (manual)", 20),
        ("Notas", 25), ("Referencia original", 50),
    ]

    # Encabezado
    ws_all.row_dimensions[1].height = 30
    for col_idx, (col_name, col_width) in enumerate(columns, 1):
        cell = ws_all.cell(row=1, column=col_idx, value=col_name)
        style_header(cell)
        ws_all.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws_all.freeze_panes = "A2"

    # Datos
    for r in records:
        row_num = r["ref_number"] + 1
        study_type = r.get("study_type_auto", "otro/no_clasificado")
        row_color = STUDY_TYPE_COLORS.get(study_type, COLORS["other"])

        values = [
            r.get("ref_number", ""),
            r.get("authors", ""),
            r.get("year", ""),
            r.get("title", ""),
            r.get("journal", ""),
            r.get("pmid", ""),
            r.get("doi", ""),
            r.get("pubmed_url", ""),
            r.get("doi_url", ""),
            r.get("study_type_auto", ""),
            r.get("study_type", ""),  # campo para corrección manual
            r.get("notes", ""),
            r.get("ref_raw", ""),
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws_all.cell(row=row_num, column=col_idx, value=val)
            style_cell(cell, row_color=row_color, wrap=(col_idx in [4, 13]))
            # Hipervínculos
            if col_idx == 8 and val:
                cell.hyperlink = val
                cell.font = Font(name="Arial", size=9, color="0563C1", underline="single")
            if col_idx == 9 and val:
                cell.hyperlink = val
                cell.font = Font(name="Arial", size=9, color="0563C1", underline="single")

        add_thin_border(ws_all, row_num, 1, len(columns))
        ws_all.row_dimensions[row_num].height = 40

    # ── Hoja 2: Solo ECAs primarios ─────────────────────────────────
    ws_rct = wb.create_sheet("ECAs primarios")
    rct_records = [r for r in records if r.get("study_type_auto") == "RCT_primario"]
    _write_rct_sheet(ws_rct, rct_records)

    # ── Hoja 3: Resumen por tipo ────────────────────────────────────
    ws_sum = wb.create_sheet("Resumen")
    _write_summary_sheet(ws_sum, records)

    # ── Hoja 4: Instrucciones ───────────────────────────────────────
    ws_help = wb.create_sheet("Instrucciones")
    _write_instructions_sheet(ws_help)

    wb.save(output_path)
    print(f"[OK] Excel guardado: {output_path}")


def _write_rct_sheet(ws, records):
    ws.title = "ECAs primarios"
    cols = [
        ("N°", 5), ("Autores", 30), ("Año", 6), ("Título", 50),
        ("Revista", 25), ("PMID", 12), ("DOI", 30),
        ("URL PubMed", 35), ("URL DOI", 35), ("Notas", 30),
    ]
    ws.row_dimensions[1].height = 30
    for col_idx, (col_name, col_width) in enumerate(cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        style_header(cell, bg_color="375623")
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    ws.freeze_panes = "A2"

    for row_num, r in enumerate(records, 2):
        values = [
            r.get("ref_number"), r.get("authors"), r.get("year"),
            r.get("title"), r.get("journal"), r.get("pmid"), r.get("doi"),
            r.get("pubmed_url"), r.get("doi_url"), r.get("notes"),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            style_cell(cell, row_color=COLORS["rct_primary"], wrap=(col_idx == 4))
            if col_idx == 8 and val:
                cell.hyperlink = val
                cell.font = Font(name="Arial", size=9, color="0563C1", underline="single")
            if col_idx == 9 and val:
                cell.hyperlink = val
                cell.font = Font(name="Arial", size=9, color="0563C1", underline="single")
        ws.row_dimensions[row_num].height = 40


def _write_summary_sheet(ws, records):
    ws.title = "Resumen"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15

    headers = ["Tipo de estudio", "N referencias", "% del total"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        style_header(cell)

    type_counts = {}
    for r in records:
        t = r.get("study_type_auto", "otro/no_clasificado")
        type_counts[t] = type_counts.get(t, 0) + 1

    total = len(records)
    for row_idx, (study_type, count) in enumerate(sorted(type_counts.items()), 2):
        pct = count / total * 100 if total else 0
        row_color = STUDY_TYPE_COLORS.get(study_type, COLORS["other"])
        ws.cell(row=row_idx, column=1, value=study_type).fill = PatternFill("solid", start_color=row_color)
        ws.cell(row=row_idx, column=2, value=count)
        ws.cell(row=row_idx, column=3, value=f"{pct:.1f}%")

    # Total
    total_row = len(type_counts) + 2
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True, name="Arial")
    ws.cell(row=total_row, column=2, value=total).font = Font(bold=True, name="Arial")
    ws.cell(row=total_row, column=3, value="100%").font = Font(bold=True, name="Arial")

    # Leyenda de colores
    ws.cell(row=total_row + 2, column=1, value="LEYENDA DE COLORES:").font = Font(bold=True, name="Arial", size=9)
    for idx, (st, color) in enumerate(STUDY_TYPE_COLORS.items(), total_row + 3):
        cell = ws.cell(row=idx, column=1, value=st)
        cell.fill = PatternFill("solid", start_color=color)
        cell.font = Font(name="Arial", size=9)


def _write_instructions_sheet(ws):
    ws.title = "Instrucciones"
    ws.column_dimensions["A"].width = 80

    lines = [
        ("GUÍA DE USO DEL PIPELINE DE REFERENCIAS", True, COLORS["header_bg"], "FFFFFF"),
        ("", False, None, None),
        ("MÓDULO 1 — Extracción del PDF", True, COLORS["subheader"], "000000"),
        ("El script extrae automáticamente la sección 'References' del PDF de la guía.", False, None, None),
        ("Detecta referencias numeradas (1. Autor... o 1 Autor...) y las limpia.", False, None, None),
        ("", False, None, None),
        ("MÓDULO 2 — Enriquecimiento de metadatos", True, COLORS["subheader"], "000000"),
        ("Cada referencia se busca en CrossRef (texto libre) y PubMed (query por título+año).", False, None, None),
        ("Se extraen: PMID, DOI, título, autores, año, revista.", False, None, None),
        ("Se añaden URLs clicables a PubMed y DOI.", False, None, None),
        ("NOTA: Sin API key de NCBI el límite es 3 req/s. El script respeta este límite.", False, None, None),
        ("", False, None, None),
        ("MÓDULO 3 — Clasificación automática", True, COLORS["subheader"], "000000"),
        ("Se clasifica cada referencia por palabras clave en el título:", False, None, None),
        ("  • RCT_primario: ensayo clínico aleatorizado, publicación principal", False, None, None),
        ("  • RCT_secundario: subanálisis, post-hoc, subgrupos", False, None, None),
        ("  • meta-analisis: revisión sistemática, meta-análisis, pooled analysis", False, None, None),
        ("  • registro_observacional: registro, cohorte, observacional, retrospectivo", False, None, None),
        ("  • guia_clinica: guideline, consensus statement", False, None, None),
        ("  • otro/no_clasificado: no coincide con ningún patrón", False, None, None),
        ("La columna 'Tipo (manual)' permite correcciones manuales.", False, None, None),
        ("", False, None, None),
        ("MÓDULO 4 — Excel estructurado", True, COLORS["subheader"], "000000"),
        ("Hoja 'Todas las referencias': base de datos completa con código de colores por tipo.", False, None, None),
        ("Hoja 'ECAs primarios': solo los ECAs primarios identificados.", False, None, None),
        ("Hoja 'Resumen': tabla de frecuencias por tipo de estudio.", False, None, None),
        ("", False, None, None),
        ("USO EN LÍNEA DE COMANDOS", True, COLORS["subheader"], "000000"),
        ("  python guideline_pipeline.py guia.pdf output.xlsx", False, None, None),
        ("  python guideline_pipeline.py guia.pdf              (usa 'references_db.xlsx' por defecto)", False, None, None),
        ("", False, None, None),
        ("PARA GUÍAS CON API KEY DE NCBI (>3 req/s)", True, COLORS["subheader"], "000000"),
        ("  Añadir al entorno: export NCBI_API_KEY=tu_clave", False, None, None),
        ("  Obtener gratis en: https://www.ncbi.nlm.nih.gov/account/", False, None, None),
    ]

    for row_idx, (text, bold, bg, fg) in enumerate(lines, 1):
        cell = ws.cell(row=row_idx, column=1, value=text)
        cell.font = Font(
            bold=bold, name="Arial", size=10,
            color=fg if fg else "000000"
        )
        if bg:
            cell.fill = PatternFill("solid", start_color=bg)
        cell.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row_idx].height = 18


# ─────────────────────────────────────────────
# PIPELINE PRINCIPAL
# ─────────────────────────────────────────────

def run_pipeline(pdf_path: str, output_path: str = "references_db.xlsx"):
    print(f"\n{'='*60}")
    print(f"  PIPELINE DE REFERENCIAS - GUÍAS CLÍNICAS")
    print(f"{'='*60}")
    print(f"  PDF: {pdf_path}")
    print(f"  Output: {output_path}\n")

    # PASO 1: Extraer referencias
    print("[PASO 1] Extrayendo referencias del PDF...")
    raw_refs = extract_references_from_pdf(pdf_path)

    if not raw_refs:
        print("[ERROR] No se encontraron referencias. Verificar formato del PDF.")
        return

    # PASO 2 + 3: Enriquecer y clasificar
    print(f"\n[PASO 2+3] Enriqueciendo {len(raw_refs)} referencias con PubMed + CrossRef...")
    print("  (Esto puede tardar varios minutos para guías con muchas referencias)\n")

    records = []
    for idx, ref_text in enumerate(raw_refs, 1):
        print(f"  [{idx}/{len(raw_refs)}] {ref_text[:80]}...", end="\r")
        record = enrich_reference(ref_text, idx)
        record["study_type_auto"] = classify_reference(record)
        records.append(record)

    print(f"\n\n[INFO] Clasificación:")
    type_counts = {}
    for r in records:
        t = r["study_type_auto"]
        type_counts[t] = type_counts.get(t, 0) + 1
    for t, c in sorted(type_counts.items()):
        print(f"  {t}: {c}")

    # PASO 4: Exportar
    print(f"\n[PASO 4] Exportando a Excel: {output_path}")
    export_to_excel(records, output_path)

    print(f"\n{'='*60}")
    print(f"  COMPLETADO: {len(records)} referencias procesadas")
    print(f"  ECAs primarios identificados: {type_counts.get('RCT_primario', 0)}")
    print(f"{'='*60}\n")

    return records


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("USO: python guideline_pipeline.py <ruta_pdf> [output.xlsx]")
        sys.exit(1)

    pdf_file = sys.argv[1]
    out_file = sys.argv[2] if len(sys.argv) > 2 else "references_db.xlsx"
    run_pipeline(pdf_file, out_file)
